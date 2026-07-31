import json
import os
import sys
sys.path.insert(0, '.')

from collections import defaultdict
import openpyxl

v5_excel_path = r"c:\Users\ggvfj\Downloads\All Projects\Time_Table\time_table\ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx"
test5_excel_path = r"c:\Users\ggvfj\Downloads\All Projects\Time_Table\data\test_outputs\Test5_FullDepartment_44Sections_Master.xlsx"
v5_clean_json = r"c:\Users\ggvfj\Downloads\All Projects\Time_Table\data\seed\v5_faculty_assignments_clean.json"

with open(v5_clean_json, "r", encoding="utf-8") as f:
    v5_ground_truth = json.load(f)

# Parse Test 5 generated Excel output
test5_wb = openpyxl.load_workbook(test5_excel_path, data_only=True)

test5_faculty_workload = defaultdict(lambda: {
    "weekly_hours": 0,
    "daily_hours": defaultdict(int),
    "subjects": set(),
    "sections": set()
})

test5_assignments = []

for sheet_name in test5_wb.sheetnames:
    ws = test5_wb[sheet_name]
    # In Test 5 output, section sheets contain timetable grid + legend below
    # Let's extract entries from the grid (rows 5 to 10, cols 2 to 11)
    DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
    PERIOD_COLS = [(1, 2), (2, 3), (3, 5), (4, 6), (5, 7), (6, 9), (7, 10), (8, 11)]

    # Parse legend at bottom of sheet to see how faculty were assigned
    legend_start_row = 13
    for r in range(legend_start_row, ws.max_row + 1):
        subj_cell = ws.cell(row=r, column=1).value
        fac_cell = ws.cell(row=r, column=4).value or ws.cell(row=r, column=3).value
        if subj_cell and fac_cell:
            s_str = str(subj_cell).strip()
            f_str = str(fac_cell).strip()
            if s_str and f_str and f_str != "None":
                test5_assignments.append({
                    "section": sheet_name,
                    "subject": s_str,
                    "faculty": f_str
                })

print(f"Parsed {len(test5_assignments)} section-subject-faculty assignments from Test 5 output.")

# Compare Test 5 assignments vs V5 Ground Truth
misaligned_subject_assignments = []
unassigned_or_synthetic_facs = []
workload_discrepancies = []

for item in test5_assignments:
    sec = item["section"]
    subj = item["subject"]
    fac_str = item["faculty"]
    
    # Split primary and co-faculty
    facs = [f.strip() for f in fac_str.split(",") if f.strip()]
    for f in facs:
        if f in v5_ground_truth:
            gt_subjs = v5_ground_truth[f]["subjects_taught"]
            # Check if subject matches ground truth subject
            clean_s = subj.replace("(P)", "").replace("(T&P)", "").replace("(T)", "").strip()
            match_found = any(clean_s in gt_s for gt_s in gt_subjs)
            if not match_found:
                misaligned_subject_assignments.append({
                    "faculty": f,
                    "section": sec,
                    "assigned_subject": subj,
                    "ground_truth_subjects": gt_subjs
                })
        else:
            unassigned_or_synthetic_facs.append({
                "faculty": f,
                "section": sec,
                "assigned_subject": subj
            })

print("\n--- COMPARISON RESULTS ---")
print(f"1. Subject Misassignments (Faculty assigned to wrong subject): {len(misaligned_subject_assignments)}")
print(f"2. Synthetic / Unrecognized Faculty Names: {len(unassigned_or_synthetic_facs)}")

# Sample discrepancies
if misaligned_subject_assignments:
    print("\nSample Subject Misassignments in Test 5:")
    for m in misaligned_subject_assignments[:10]:
        print(f"  • {m['faculty']} assigned to '{m['assigned_subject']}' in {m['section']} (Ground Truth: {m['ground_truth_subjects']})")

if unassigned_or_synthetic_facs:
    print("\nSample Synthetic / Unrecognized Faculty:")
    for u in unassigned_or_synthetic_facs[:10]:
        print(f"  • {u['faculty']} assigned to '{u['assigned_subject']}' in {u['section']}")
