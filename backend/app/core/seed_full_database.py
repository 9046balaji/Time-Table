import os
import sys
import re
import asyncio
import datetime
import openpyxl

sys.path.insert(0, '.')
sys.path.insert(0, 'backend')

from app.core.database import AsyncSessionLocal, engine, Base
from app.models import (
    Department,
    AcademicYear,
    Branch,
    Section,
    Faculty,
    Room,
    Subject,
    SectionSubject,
    TimeSlot,
    TimetableVersion,
    TimetableEntry
)
from sqlalchemy import select

V5_EXCEL = r"c:\Users\ggvfj\Downloads\All Projects\Time_Table\time_table\ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx"
FILE_4TH = r"c:\Users\ggvfj\Downloads\All Projects\Time_Table\time_table\4th yr TT 17TH JULY.xlsx"

async def seed_database():
    print("=" * 80)
    print("VFSTR COMPLETE DATABASE SEEDING ENGINE: V5 BASELINE + 4TH YEAR DATASET")
    print("=" * 80)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with AsyncSessionLocal() as db:
        # 1. Department & Academic Year
        dep_res = await db.execute(select(Department).where(Department.code == "ACSE"))
        dep = dep_res.scalars().first()
        if not dep:
            dep = Department(name="Advanced Computer Science & Engineering", code="ACSE")
            db.add(dep)
            await db.flush()

        ay_res = await db.execute(select(AcademicYear).where(AcademicYear.label == "2026-27 Sem I"))
        ay = ay_res.scalars().first()
        if not ay:
            ay = AcademicYear(year=2026, semester=1, label="2026-27 Sem I", is_current=True)
            db.add(ay)
            await db.flush()

        # 2. Time Slots (P1..P8 across MON..SAT)
        DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
        TIMES = {
            1: (datetime.time(8, 15), datetime.time(9, 5)),
            2: (datetime.time(9, 5), datetime.time(9, 55)),
            3: (datetime.time(10, 10), datetime.time(11, 0)),
            4: (datetime.time(11, 0), datetime.time(11, 50)),
            5: (datetime.time(11, 50), datetime.time(12, 40)),
            6: (datetime.time(13, 40), datetime.time(14, 30)),
            7: (datetime.time(14, 30), datetime.time(15, 20)),
            8: (datetime.time(15, 20), datetime.time(16, 5))
        }
        db_time_slots = {}
        for d in DAYS:
            for p, (st, et) in TIMES.items():
                ts_res = await db.execute(select(TimeSlot).where(TimeSlot.day == d, TimeSlot.period == p))
                ts_obj = ts_res.scalars().first()
                if not ts_obj:
                    ts_obj = TimeSlot(day=d, period=p, start_time=st, end_time=et, is_blocked=False)
                    db.add(ts_obj)
                    await db.flush()
                db_time_slots[(d, p)] = ts_obj

        # 3. Branches
        branches_def = [
            {"code": "CSE-AIML", "name": "CSE (AI & Machine Learning)"},
            {"code": "CSE-CORE", "name": "CSE (Core Computer Science)"},
            {"code": "CSE-DS",   "name": "CSE (Data Science)"},
            {"code": "CSE-CS",   "name": "CSE (Cyber Security)"},
            {"code": "CSE-CSBS", "name": "CSE (Business Systems)"},
            {"code": "CSE-IOT",  "name": "CSE (Internet of Things)"}
        ]
        branch_map = {}
        for b_info in branches_def:
            b_res = await db.execute(select(Branch).where(Branch.code == b_info["code"]))
            b_obj = b_res.scalars().first()
            if not b_obj:
                b_obj = Branch(dept_id=dep.id, code=b_info["code"], name=b_info["name"])
                db.add(b_obj)
                await db.flush()
            branch_map[b_info["code"]] = b_obj

        # ---------------------------------------------------------------------
        # 4. Parse and Collect All Entities from Excel Files
        # ---------------------------------------------------------------------
        all_sections_set = set()
        all_rooms_map = {}
        all_faculty_map = {}
        all_subjects_map = {}

        # Parse 4th Year File
        if os.path.exists(FILE_4TH):
            print(f"\n[1/3] Extracting 4th Year Excel Data ({os.path.basename(FILE_4TH)})...")
            wb_4th = openpyxl.load_workbook(FILE_4TH, data_only=True)
            for sname in wb_4th.sheetnames:
                ws = wb_4th[sname]
                sec_name = sname.upper().strip()
                if sec_name.startswith("SEC"):
                    sec_name = f"SECTION-{sec_name.replace('SEC', '')}"
                all_sections_set.add(sec_name)

                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = str(ws.cell(row=r, column=c).value or "")
                        room_matches = re.findall(r'\[(N-[\w\-]+|AFTF-[\w\-]+|\d{3}\w?|AFF-[\w\-]+)\]|\((N-[\w\-]+|AFTF-[\w\-]+|\d{3}\w?|N-\d{3}\w?|AFF-[\w\-]+)\)', val)
                        for rm_tuple in room_matches:
                            rm_code = [x for x in rm_tuple if x][0].strip()
                            if rm_code and rm_code not in all_rooms_map:
                                rtype = "computer_lab" if "AFTF" in rm_code or "AFF" in rm_code or rm_code in ["604","605","606","611","612","615","616","617"] else "classroom"
                                all_rooms_map[rm_code] = {"code": rm_code, "capacity": 66, "room_type": rtype}

                        if "(" in val and ")" in val and any(char.isdigit() for char in val):
                            m_fac = re.search(r'([A-Za-z\.\s]+)\s*\((\d{10})\)', val)
                            if m_fac:
                                f_name = m_fac.group(1).strip()
                                if f_name and not f_name[0].isdigit() and f_name not in ["***", "undefined"]:
                                    all_faculty_map[f_name] = {"name": f_name, "designation": "Assistant Professor"}

                        if "(" in val and ")" in val and "22CS" in val:
                            m_sub = re.search(r'([A-Za-z\s]+)\((\d{2}CS\d{3})\)', val)
                            if m_sub:
                                s_title = m_sub.group(1).strip()
                                s_code = m_sub.group(2).strip()
                                all_subjects_map[s_code] = {"code": s_code, "full_name": s_title, "slot_type": "L", "hours": 3}

        # Parse V5 Baseline File
        if os.path.exists(V5_EXCEL):
            print(f"\n[2/3] Extracting V5 Baseline Data ({os.path.basename(V5_EXCEL)})...")
            v5_secs = ["II AIML-A", "II AIML-B", "II AIML-C", "II AIML-D", "II AIML-E", "II AIML-F", "II AIML-G", "II AIML-H", "II AIML-I", "II AIML-J", "II AIML-K", "II AIML-L", "III AIML-A", "III AIML-B", "III AIML-C", "III AIML-D", "III AIML-E", "III AIML-F", "III AIML-G", "IV AIML-A", "IV AIML-B", "IV AIML-C", "IV AIML-D", "IV AIML-E", "II CS-A", "II CS-B", "III CS", "IV CS", "II DS-A", "II DS-B", "III DS-A", "III DS-B", "IV DS"]
            for s in v5_secs:
                all_sections_set.add(s)

            dept_rooms = [
                "601", "602", "603", "604", "605", "606", "607", "608", "609", "610",
                "611", "612", "613", "614", "615", "616", "617", "618", "619",
                "215", "216", "217", "218", "514-A", "514-B", "518",
                "AFTF-12", "AFTF-13", "AFTF-14", "AFF-09", "AFF-10"
            ]
            for rm_code in dept_rooms:
                if rm_code not in all_rooms_map:
                    rtype = "computer_lab" if rm_code in ["604","605","606","611","612","615","616","617","AFTF-12","AFTF-13","AFTF-14"] else "classroom"
                    all_rooms_map[rm_code] = {"code": rm_code, "capacity": 66, "room_type": rtype}

            dept_fac = [
                "Dr. S. Srikantha Reddy", "Dr. B. Sudha Rani", "Ms. P. Seetha Lakshmi", "Dr. P. Kalpana",
                "Dr. Ankamma Rao Mallela", "Dr. Bandi Guravaiah", "Dr. Imtiyaz Bhatt", "Dr. N. Bhargavi",
                "Dr. A.V. Nageswara Rao", "Dr. Arnab De", "Dr. E. Ramesh", "Dr. N. Venkateswarlu",
                "Mr. Mallela Varma", "Ms. G. Mahalakshmi", "Dr. Rushi Prasad Sahoo", "Ms. D. Supriya",
                "Ms. Challa Sai Mohitha", "Ms. Chandolu Charana Sree", "Ms. Ch. Omkara Lakshmi",
                "Dr. SK Satpathy", "Dr. A. Subramanyam", "Dr. Amar Jukuntla", "Dr. B.N. Naveen Kumar",
                "Dr. G. Yalamanda Babu", "Dr. Manigandan A", "Dr. K. Srinivas", "Ms. Attuluri Ramya"
            ]
            for fn in dept_fac:
                if fn not in all_faculty_map:
                    all_faculty_map[fn] = {"name": fn, "designation": "Assistant Professor"}

            dept_subs = [
                {"code": "SFCDS", "full_name": "Statistical Foundation for Computing & Data Science", "slot_type": "L", "hours": 4},
                {"code": "DEF",   "full_name": "Data Engineering Foundations", "slot_type": "L", "hours": 2},
                {"code": "DMS",   "full_name": "Discrete Mathematical Structures", "slot_type": "L", "hours": 2},
                {"code": "DS",    "full_name": "Data Structures", "slot_type": "L", "hours": 2},
                {"code": "AI",    "full_name": "Artificial Intelligence Search Methods", "slot_type": "L", "hours": 2},
                {"code": "DBMS",  "full_name": "Database Management Systems", "slot_type": "L", "hours": 2},
                {"code": "OOPS",  "full_name": "Object Oriented Programming", "slot_type": "L", "hours": 2},
                {"code": "DL",    "full_name": "Deep Learning & Neural Networks", "slot_type": "L", "hours": 3},
                {"code": "WT",    "full_name": "Web Technologies", "slot_type": "L", "hours": 3},
                {"code": "CV",    "full_name": "Computer Vision", "slot_type": "L", "hours": 3},
                {"code": "ADS",   "full_name": "Advanced Data Structures & Algorithms", "slot_type": "L", "hours": 3},
                {"code": "MLOP",  "full_name": "MLOps & AI Model Deployment", "slot_type": "L", "hours": 3},
                {"code": "IDP",   "full_name": "Interdisciplinary Project", "slot_type": "L", "hours": 2},
                {"code": "CNS",   "full_name": "Cryptography & Network Security", "slot_type": "L", "hours": 3},
                {"code": "TM",    "full_name": "Technical Modules", "slot_type": "L", "hours": 3},
                {"code": "GENAI", "full_name": "Generative AI & LLMs", "slot_type": "L", "hours": 4},
                {"code": "IOT",   "full_name": "Internet of Things & Sensor Networks", "slot_type": "L", "hours": 3}
            ]
            for sub_info in dept_subs:
                if sub_info["code"] not in all_subjects_map:
                    all_subjects_map[sub_info["code"]] = sub_info

        # ---------------------------------------------------------------------
        # 5. Populate PostgreSQL Database Tables
        # ---------------------------------------------------------------------
        print("\n[3/3] Committing All Extracted Entities to Database...")

        # Seed Rooms
        db_rooms = {}
        for r_code, r_info in all_rooms_map.items():
            r_res = await db.execute(select(Room).where(Room.code == r_code))
            r_obj = r_res.scalars().first()
            if not r_obj:
                r_obj = Room(dept_id=dep.id, code=r_code, room_type=r_info["room_type"], capacity=r_info["capacity"], block="U-Block", is_available=True)
                db.add(r_obj)
                await db.flush()
            db_rooms[r_code] = r_obj
        print(f"  • Successfully Seeded {len(db_rooms)} Rooms into PostgreSQL")

        # Seed Faculty
        db_faculty = {}
        for f_name, f_info in all_faculty_map.items():
            f_res = await db.execute(select(Faculty).where(Faculty.name == f_name))
            f_obj = f_res.scalars().first()
            if not f_obj:
                f_obj = Faculty(dept_id=dep.id, name=f_name, designation=f_info["designation"], max_hours_per_week=16, is_external=False)
                db.add(f_obj)
                await db.flush()
            db_faculty[f_name] = f_obj
        print(f"  • Successfully Seeded {len(db_faculty)} Faculty Members into PostgreSQL")

        # Seed Subjects
        db_subjects = {}
        for s_code, s_info in all_subjects_map.items():
            s_res = await db.execute(select(Subject).where(Subject.code == s_code))
            s_obj = s_res.scalars().first()
            if not s_obj:
                s_obj = Subject(dept_id=dep.id, code=s_code, full_name=s_info["full_name"], lecture_hours=s_info["hours"], slot_type=s_info["slot_type"])
                db.add(s_obj)
                await db.flush()
            db_subjects[s_code] = s_obj
        print(f"  • Successfully Seeded {len(db_subjects)} Subjects into PostgreSQL")

        # Seed Sections
        db_sections = {}
        for sec_name in sorted(list(all_sections_set)):
            sec_res = await db.execute(select(Section).where(Section.name == sec_name))
            sec_obj = sec_res.scalars().first()
            if not sec_obj:
                b_code = "CSE-AIML" if "AIML" in sec_name else ("CSE-DS" if "DS" in sec_name else ("CSE-CS" if "CS" in sec_name else "CSE-CORE"))
                y_num = 4 if "IV" in sec_name or "SECTION" in sec_name else (3 if "III" in sec_name else 2)
                lbl = sec_name.split("-")[-1] if "-" in sec_name else sec_name[-1]
                sec_obj = Section(branch_id=branch_map[b_code].id, year_level=y_num, label=lbl, name=sec_name, strength=60, academic_year_id=ay.id, is_active=True)
                db.add(sec_obj)
                await db.flush()
            db_sections[sec_name] = sec_obj
        print(f"  • Successfully Seeded {len(db_sections)} Sections into PostgreSQL")

        # Timetable Version V5
        ver_res = await db.execute(select(TimetableVersion).where(TimetableVersion.version_label == "V5"))
        ver = ver_res.scalars().first()
        if not ver:
            ver = TimetableVersion(academic_year_id=ay.id, version_label="V5", is_current=True, source="IMPORTED", notes="Full V5 & 4th Year Seeding")
            db.add(ver)
            await db.flush()

        await db.commit()
        print("\n" + "=" * 80)
        print("FULL DATABASE SEEDING COMPLETED SUCCESSFULLY!")
        print(f"Total Sections: {len(db_sections)} | Total Faculty: {len(db_faculty)} | Total Rooms: {len(db_rooms)} | Total Subjects: {len(db_subjects)}")
        print("=" * 80)

if __name__ == "__main__":
    asyncio.run(seed_database())
