import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any, Optional

from parser.excel_parser import ExcelTimetableParser, resolve_v5_path


COHORT_GROUPS = {
    "II_AIML": {
        "label": "B.Tech II Year AIML (Sections A-L)",
        "sections": ["II AIML-A", "II AIML-B", "II AIML-C", "II AIML-D", "II AIML-E", "II AIML-F", "II AIML-G", "II AIML-H", "II AIML-I", "II AIML-J", "II AIML-K", "II AIML-L"]
    },
    "III_AIML": {
        "label": "B.Tech III Year AIML (Sections A-G)",
        "sections": ["III AIML-A", "III AIML-B", "III AIML-C", "III AIML-D", "III AIML-E", "III AIML-F", "III AIML-G"]
    },
    "IV_AIML": {
        "label": "B.Tech IV Year AIML (Sections A-E)",
        "sections": ["IV AIML-A", "IV AIML-B", "IV AIML-C", "IV AIML-D", "IV AIML-E"]
    },
    "CS_DS": {
        "label": "B.Tech CS & DS (All Years)",
        "sections": ["II CS-A", "II CS-B", "III CS", "IV CS", "II DS-A", "II DS-B", "III DS-A", "III DS-B", "IV DS"]
    },
    "CSBS_IOT": {
        "label": "B.Tech CSBS & IOT (All Years)",
        "sections": ["II CSBS", "III CSBS", "IV - CSBS", "II IOT", "III IOT"]
    },
    "SPECIAL_PG": {
        "label": "Special Programs & Minor/Honors",
        "sections": ["II BS(DS)", "III BS(DS)", "II MSC (DS)", "MINORS/HONORS"]
    }
}


class ExcelTimetableExporter:
    PERIOD_HEADERS = ["Period", "1", "2", "09:55 - 10:10", "3", "4", "5", "12:40 - 1:40", "6", "7", "8"]
    TIME_SLOT_SUBHEADERS = ["Day/Hour", "8:15-9:05", "9:05-09:55", "09:55 - 10:10", "10:10-11:00", "11:00-11:50", "11:50-12:40", "12:40 - 1:40", "1:40-2:30", "2:30-3:20", "3:20-4:05"]
    DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]

    def _write_section_block(self, ws: Any, start_row: int, sec_name: str, slots: List[Dict[str, Any]]) -> int:
        """Render complete weekly section timetable block at start_row and return next start row after 3 blank spacers."""
        header_font = Font(name="Calibri", size=10, bold=True, color="000000")
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        banner_font = Font(name="Calibri", size=11, bold=True, color="000000")
        banner_fill = PatternFill(start_color="C084FC", end_color="C084FC", fill_type="solid") # Soft Purple Banner

        subj_font = Font(name="Calibri", size=9, bold=True, color="000000")
        bold_font = Font(name="Calibri", size=10, bold=True)
        cell_font = Font(name="Calibri", size=9, bold=False)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row start_row: Academic year
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=11)
        ws.cell(row=start_row, column=1, value="Academic year 2026- 27 (I Semester)").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=start_row, column=1).alignment = center_align

        # Row start_row+2: Purple Section Name Banner
        banner_row = start_row + 2
        ws.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=11)
        b_cell = ws.cell(row=banner_row, column=1, value=sec_name)
        b_cell.font = banner_font
        b_cell.fill = banner_fill
        b_cell.alignment = center_align

        # Row start_row+3: Period Headers
        p_header_row = start_row + 3
        for col_idx, text in enumerate(self.PERIOD_HEADERS, start=1):
            cell = ws.cell(row=p_header_row, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Row start_row+4: Time Subheaders
        t_header_row = start_row + 4
        for col_idx, text in enumerate(self.TIME_SLOT_SUBHEADERS, start=1):
            cell = ws.cell(row=t_header_row, column=col_idx, value=text)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align

        grid_start_row = start_row + 5
        grid_end_row = start_row + 10

        # Merged BREAK Column (Col D / Col 4)
        ws.merge_cells(start_row=grid_start_row, start_column=4, end_row=grid_end_row, end_column=4)
        break_cell = ws.cell(row=grid_start_row, column=4, value="B\nR\nE\nA\nK")
        break_cell.font = Font(name="Calibri", size=10, bold=True)
        break_cell.alignment = center_align
        for r in range(grid_start_row, grid_end_row + 1):
            ws.cell(row=r, column=4).border = thin_border

        # Merged LUNCH Column (Col H / Col 8)
        ws.merge_cells(start_row=grid_start_row, start_column=8, end_row=grid_end_row, end_column=8)
        lunch_cell = ws.cell(row=grid_start_row, column=8, value="L\nU\nN\nC\nH")
        lunch_cell.font = Font(name="Calibri", size=10, bold=True)
        lunch_cell.alignment = center_align
        for r in range(grid_start_row, grid_end_row + 1):
            ws.cell(row=r, column=8).border = thin_border

        faculty_lecture_map = {}
        faculty_lab_map = {}

        # Rows grid_start_row..grid_end_row: Days MON..SAT
        for d_off, day in enumerate(self.DAYS):
            r_num = grid_start_row + d_off
            ws.cell(row=r_num, column=1, value=day).font = bold_font
            ws.cell(row=r_num, column=1).alignment = center_align
            ws.cell(row=r_num, column=1).border = thin_border

            period_mapping = [
                (1, 2), (2, 3), (3, 5), (4, 6), (5, 7), (6, 9), (7, 10), (8, 11)
            ]

            for p_num, col_idx in period_mapping:
                cell = ws.cell(row=r_num, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

                matching = [s for s in slots if s.get("section") == sec_name and s.get("day") == day and s.get("period") == p_num]
                if matching:
                    m = matching[0]
                    subj = m.get("subject", "")
                    room = m.get("room", "")
                    fac = m.get("faculty", "")

                    cell.value = f"{subj}\n{room}" if room else subj
                    cell.font = subj_font

                    if subj and fac:
                        if "(P)" in subj or "(T&P)" in subj or "Lab" in subj:
                            faculty_lab_map[subj] = fac
                        else:
                            faculty_lecture_map[subj] = fac
                else:
                    cell.value = ""

        # 2-Column Faculty Allocation Legend Table
        start_leg_row = grid_end_row + 2
        all_subjs = sorted(list(set(list(faculty_lecture_map.keys()) + list(faculty_lab_map.keys()))))
        last_leg_row = start_leg_row

        if all_subjs:
            for offset, subj_code in enumerate(all_subjs):
                s_idx = start_leg_row + offset
                last_leg_row = s_idx

                # Left Column (A..E): Lecture Faculty
                ws.merge_cells(start_row=s_idx, start_column=1, end_row=s_idx, end_column=5)
                lec_fac = faculty_lecture_map.get(subj_code, "")
                lec_text = f"{subj_code}(L): {lec_fac}" if lec_fac else f"{subj_code}: Course Allocation"
                ws.cell(row=s_idx, column=1, value=lec_text).font = cell_font
                for c in range(1, 6):
                    ws.cell(row=s_idx, column=c).border = thin_border

                # Right Column (F..K): Practical/Tutorial Instructors
                ws.merge_cells(start_row=s_idx, start_column=6, end_row=s_idx, end_column=11)
                lab_fac = faculty_lab_map.get(subj_code, "")
                lab_text = f"{subj_code}(P): {lab_fac}" if lab_fac else f"{subj_code}(T&P): Faculty Team"
                ws.cell(row=s_idx, column=6, value=lab_text).font = cell_font
                for c in range(6, 12):
                    ws.cell(row=s_idx, column=c).border = thin_border
        else:
            last_leg_row = start_leg_row
            ws.cell(row=last_leg_row, column=1, value="All section slots assigned to department instructors.").font = cell_font

        ws.column_dimensions['A'].width = 12
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[c].width = 15

        # Return next start row after 3 blank spacer rows (Matching Screenshot 2026-07-23 213129.png)
        return last_leg_row + 4

    def export_timetable(self, timetable_data: Optional[Dict[str, Any]] = None) -> bytes:
        timetable_data = timetable_data or {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        sections = timetable_data.get("sections")
        slots = timetable_data.get("slots") or []

        # Parse V5 if slots or sections empty
        if not sections or not slots:
            parser = ExcelTimetableParser()
            try:
                v5_path = resolve_v5_path()
                parsed_res = parser.parse_file(v5_path)
                if not sections:
                    sections = [{"name": sname} for sname in parsed_res.sections.keys()]
                if not slots:
                    slots = [
                        {
                            "section": s.section,
                            "day": s.day,
                            "period": s.period,
                            "subject": s.subject_code,
                            "room": s.room or "",
                            "faculty": ", ".join(s.faculty_list) if s.faculty_list else ""
                        }
                        for s in parsed_res.raw_entries
                    ]
            except Exception as e:
                print(f"[ExcelExporter Warning] Could not parse V5 file: {e}")
                if not sections:
                    sections = [{"name": f"Section {i+1}"} for i in range(44)]
                if not slots:
                    slots = []

        # 1. Master Department Sheet (All Sections Stacked Vertically with Spacers)
        ws_master = wb.create_sheet(title="Master Department View")
        curr_row = 2
        for sec in sections:
            sec_name = sec.get("name", "Section")
            curr_row = self._write_section_block(ws_master, curr_row, sec_name, slots)

        # 2. Individual Section Worksheets
        for sec in sections:
            sec_name = sec.get("name", "Section")
            sheet_title = sec_name[:31]
            ws = wb.create_sheet(title=sheet_title)
            self._write_section_block(ws, 2, sec_name, slots)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_cohort_excel(self, cohort_key: str, timetable_data: Optional[Dict[str, Any]] = None) -> bytes:
        """Export cohort-specific Excel workbook with Sections Vertical Stacked on Sheet 1 + individual section tabs."""
        timetable_data = timetable_data or {}
        cohort_info = COHORT_GROUPS.get(cohort_key, COHORT_GROUPS["II_AIML"])
        target_sections = cohort_info["sections"]
        cohort_label = cohort_info["label"]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        slots = timetable_data.get("slots") or []

        if not slots:
            parser = ExcelTimetableParser()
            try:
                v5_path = resolve_v5_path()
                parsed_res = parser.parse_file(v5_path)
                slots = [
                    {
                        "section": s.section,
                        "day": s.day,
                        "period": s.period,
                        "subject": s.subject_code,
                        "room": s.room or "",
                        "faculty": ", ".join(s.faculty_list) if s.faculty_list else ""
                    }
                    for s in parsed_res.raw_entries
                ]
            except Exception as e:
                print(f"[ExcelExporter Cohort Warning] {e}")

        # TAB 1: COHORT STACKED MASTER VIEW (Sections Stacked Vertically with 3 Spacers)
        ws_cohort = wb.create_sheet(title="Cohort Stacked Master View")
        curr_row = 2
        for sec_name in target_sections:
            curr_row = self._write_section_block(ws_cohort, curr_row, sec_name, slots)

        # TABS 2+: Individual Section Worksheets
        for sec_name in target_sections:
            ws = wb.create_sheet(title=sec_name[:31])
            self._write_section_block(ws, 2, sec_name, slots)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_minors_honors_excel(self, timetable_data: Optional[Dict[str, Any]] = None) -> bytes:
        """Generate Minors/Honors Department Master Allocation Sheet matching Screenshot 2026-07-23 214040.png."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MINORS_HONORS"

        yellow_banner_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid") # Bright Yellow
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        title_font = Font(name="Calibri", size=11, bold=True, color="000000")
        red_header_font = Font(name="Calibri", size=10, bold=True, color="DC2626")
        cell_font = Font(name="Calibri", size=9, bold=False)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        branches_data = [
            ("AIML", [
                ("III", "I", "Minor", "CS301", "Digital Image Processing", "A", "NB518", "", "***", "", "3-0-2-4"),
                ("III", "I", "Honors", "CS302", "Web and Sequence Data Mining", "A", "NB-218", "", "***", "", "3-0-2-4"),
                ("III", "I", "Honors", "CS303", "Deep Learning", "B", "NB-614", "", "***", "", "3-0-2-4"),
                ("IV", "I", "Honors", "CS401", "Cloud Computing for Machine Learning", "A", "NB-501", "***", "", "", "0-4-4-4")
            ]),
            ("CS", [
                ("III", "I", "Honors", "CS305", "Vulnerability Assessment", "A", "514-B", "A. Hruday Raj", "***", "A. Hruday Raj", "3-0-2-4"),
                ("IV", "I", "Honors", "CS405", "Security Audit", "A", "514-B", "Ms. Attuluri Ramy", "***", "Ms. Attuluri Ramy", "3-0-2-4")
            ]),
            ("CSBS", [
                ("III", "I", "Honors", "CB301", "Image Processing & Pattern Recognition", "A", "402", "***", "", "", "3-0-2-3"),
                ("IV", "I", "Honors", "CB401", "Business Analytics & Decision Systems", "A", "502", "***", "", "", "0-4-4-4")
            ]),
            ("DS", [
                ("III", "I", "Honors", "DS301", "Web and Sequence Data Mining", "A", "514-A", "", "***", "", "3-0-2-4"),
                ("IV", "I", "Honors", "DS401", "Business Analytics & Decision Systems", "A", "501", "***", "", "", "0-4-4-4")
            ]),
            ("IoT", [
                ("III", "I", "Honors", "IT301", "Machine Learning for IoT Systems", "A", "NB-514", "", "", "", "3-0-2-4")
            ])
        ]

        curr_row = 1
        headers = ["Year", "Semester", "Department Elective/Minor/Honor", "Course Code", "Course Name", "Section", "Room No", "Lecture", "Tutorial", "Practical", "L-T-P-C"]

        for b_name, rows_list in branches_data:
            # Yellow Branch Header
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=11)
            b_cell = ws.cell(row=curr_row, column=1, value=b_name)
            b_cell.font = title_font
            b_cell.fill = yellow_banner_fill
            b_cell.alignment = center_align
            curr_row += 1

            # Subheaders
            for col_idx, text in enumerate(headers, start=1):
                c = ws.cell(row=curr_row, column=col_idx, value=text)
                c.font = red_header_font
                c.fill = header_fill
                c.alignment = center_align
                c.border = thin_border
            curr_row += 1

            # Rows
            for r_tuple in rows_list:
                for col_idx, val in enumerate(r_tuple, start=1):
                    c = ws.cell(row=curr_row, column=col_idx, value=val)
                    c.font = cell_font
                    c.alignment = center_align
                    c.border = thin_border
                curr_row += 1
            curr_row += 1  # Spacer row

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
