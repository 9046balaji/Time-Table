import openpyxl
import re
import os
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Any, Set

from backend.parser.normalizer import (
    normalize_faculty_name,
    normalize_room_code,
    normalize_subject_code,
    normalize_section_name,
    SUBJECT_TITLE_TO_CODE,
)

V5_FILE_PATH = "time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx"
V3_FILE_PATH = "time_table/ACSE TIMETABLE (V3)  - W.e.f 13-7-2026.xlsx"


def resolve_v5_path(v_name: str = "V5") -> str:
    return resolve_version_path(v_name)


def resolve_version_path(v_name: str = "V5") -> str:
    if str(v_name).upper() in ("3", "V3"):
        candidates = [
            V3_FILE_PATH,
            "data/ACSE_TIMETABLE_V3.xlsx",
            "../time_table/ACSE TIMETABLE (V3)  - W.e.f 13-7-2026.xlsx",
            "/app/time_table/ACSE TIMETABLE (V3)  - W.e.f 13-7-2026.xlsx"
        ]
    else:
        candidates = [
            V5_FILE_PATH,
            "data/ACSE_TIMETABLE_V5.xlsx",
            "../time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx",
            "/app/time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx"
        ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[0]



@dataclass
class ParsedSlot:
    section: str
    day: str
    period: int
    subject_code: str
    room: str
    subject_type: str = "L"  # L, T, P, LIBRARY, BREAK, LUNCH, MINORHONOR, PROJECT
    faculty_list: List[str] = field(default_factory=list)
    raw_cell: str = ""
    sheet_name: str = ""
    time_window: str = ""
    class_teacher: str = ""


@dataclass
class ParsedResult:
    total_sections: int = 0
    total_slots: int = 0
    sections: Dict[str, List[ParsedSlot]] = field(default_factory=dict)
    faculty_mappings: Dict[str, Dict[str, List[str]]] = field(default_factory=dict)
    raw_entries: List[ParsedSlot] = field(default_factory=list)


class ExcelTimetableParser:
    KNOWN_SECTION_PREFIXES = (
        "II AIML", "III AIML", "IV AIML",
        "II CS", "III CS", "IV CS",
        "II DS", "III DS", "IV DS",
        "II CSBS", "III CSBS", "IV CSBS", "IV - CSBS",
        "II IOT", "III IOT", "IV IOT",
        "I BS(DS)", "II BS(DS)", "III BS(DS)",
        "I MSC", "II MSC", "I MTECH", "II MTECH",
        "MINOR", "HONOR", "MINORHONORS", "MINORS/HONORS"
    )

    DAY_NORM_MAP = {
        "MON": "MON", "MONDAY": "MON",
        "TUE": "TUE", "TUESDAY": "TUE",
        "WED": "WED", "WEDNESDAY": "WED",
        "THU": "THU", "THURSDAY": "THU",
        "FRI": "FRI", "FRIDAY": "FRI",
        "SAT": "SAT", "SATURDAY": "SAT",
    }

    COL_PERIOD_MAP = {
        2: 1, 3: 2,
        5: 3, 6: 4, 7: 5,
        9: 6, 10: 7, 11: 8
    }

    def parse_file(self, file_path: str, max_sections: Optional[int] = None) -> ParsedResult:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = ParsedResult()

        all_slots: List[ParsedSlot] = []
        sections_dict: Dict[str, List[ParsedSlot]] = {}
        faculty_map: Dict[str, Dict[str, List[str]]] = {}

        # Auto-detect 4th year per-section format (e.g., sheets named sec1..sec19 or SECTION- headers)
        is_per_section_format = any(
            name.lower().startswith("sec") and name[3:].isdigit() for name in wb.sheetnames
        )

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if is_per_section_format:
                self._parse_fourth_year_sheet(sheet, sheet_name, all_slots, sections_dict, faculty_map)
            else:
                self._parse_sheet(sheet, sheet_name, all_slots, sections_dict, faculty_map)

        # Filter out empty dummy section keys if any
        sections_dict = {k: v for k, v in sections_dict.items() if len(v) > 0}

        # Focus on at most max_sections if specified
        if max_sections and max_sections > 0:
            target_keys = list(sections_dict.keys())[:max_sections]
            sections_dict = {k: sections_dict[k] for k in target_keys}
            all_slots = [s for s in all_slots if s.section in sections_dict]
            faculty_map = {k: faculty_map[k] for k in target_keys if k in faculty_map}

        result.raw_entries = all_slots
        result.sections = sections_dict
        result.total_sections = len(sections_dict)
        result.total_slots = len(all_slots)
        result.faculty_mappings = faculty_map

        return result


    def _parse_fourth_year_sheet(self, sheet: Any, sheet_name: str, all_slots: List[ParsedSlot],
                                 sections_dict: Dict[str, List[ParsedSlot]],
                                 faculty_map: Dict[str, Dict[str, List[str]]]):
        sec_title = sheet_name.replace("sec", "SECTION-")
        headers: List[str] = []
        default_rooms: Dict[int, str] = {}
        class_teacher = ""

        # 1. Header & Default room scanning
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
            non_empty = [str(v).strip() for v in row_vals if v is not None and str(v).strip() != ""]
            if not non_empty:
                continue
            if "SECTION" in non_empty[0].upper():
                sec_title = non_empty[0].strip()
                continue
            if non_empty[0].lower() in ["day", "days"]:
                headers = [str(v).strip() if v is not None else "" for v in row_vals]
                for c_idx, h in enumerate(row_vals):
                    if h and isinstance(h, str):
                        rm = re.search(r'\(?(N-[\w\-]+|AFTF-[\w\-]+|\d{3}\w?)\)?', h)
                        if rm and rm.group(1) not in ["8.30", "10.50", "11.40", "12.45", "1.30", "2.20", "3.10", "4.00"]:
                            default_rooms[c_idx] = rm.group(1)
                break

        if sec_title not in sections_dict:
            sections_dict[sec_title] = []
        if sec_title not in faculty_map:
            faculty_map[sec_title] = {}

        # 2. Legend & Class Teacher scanning
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
            non_empty = [v for v in row_vals if v is not None and str(v).strip() != ""]
            if not non_empty:
                continue
            first_val = str(non_empty[0]).strip()
            if "CLASS TEACHER" in first_val.upper() and len(non_empty) > 1:
                class_teacher = str(non_empty[1]).strip()
            elif len(non_empty) >= 2 and ("22CS" in first_val or any(k in first_val for k in ["Privacy", "Big Data", "Cloud", "Machine", "Natural", "Agentic"])):
                f_list = [normalize_faculty_name(f) for f in re.split(r"[,;/&]", str(non_empty[-1])) if f.strip()]
                faculty_map[sec_title][first_val] = f_list

        # 3. Day row scanning
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
            non_empty = [str(v).strip() for v in row_vals if v is not None and str(v).strip() != ""]
            if not non_empty:
                continue
            first_val = str(non_empty[0]).strip().capitalize()
            if first_val in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
                day_norm = self.DAY_NORM_MAP.get(first_val.upper(), first_val[:3].upper())
                for c_idx in range(1, min(len(headers), len(row_vals))):
                    val = row_vals[c_idx]
                    if not val:
                        continue
                    val_str = str(val).strip()
                    if not val_str or val_str.upper() in ["BREAK", "LUNCH"]:
                        continue

                    period = c_idx
                    h_time = headers[c_idx] if c_idx < len(headers) else ""

                    rm_match = re.search(r'\[(N-[\w\-]+|AFTF-[\w\-]+|\d{3}\w?)\]|\((N-[\w\-]+|AFTF-[\w\-]+|\d{3}\w?)\)', val_str)
                    if rm_match:
                        room = rm_match.group(1) or rm_match.group(2)
                    else:
                        room = default_rooms.get(c_idx, "")

                    stype = "L"
                    if "[P]" in val_str or "(P)" in val_str or "LAB" in val_str.upper():
                        stype = "P"
                    elif "[T]" in val_str or "(T)" in val_str or "TUTORIAL" in val_str.upper():
                        stype = "T"
                    elif "PROJECT" in val_str.upper() or "SELF LEARNING" in val_str.upper():
                        stype = "SL_EL"
                    elif "MINOR" in val_str.upper():
                        stype = "MINORHONOR"

                    slot = ParsedSlot(
                        section=sec_title,
                        day=day_norm,
                        period=period,
                        subject_code=val_str,
                        room=room,
                        subject_type=stype,
                        raw_cell=val_str,
                        sheet_name=sheet_name,
                        time_window=h_time,
                        class_teacher=class_teacher
                    )

                    all_slots.append(slot)
                    sections_dict[sec_title].append(slot)


    def _parse_sheet(self, sheet: Any, sheet_name: str, all_slots: List[ParsedSlot],
                     sections_dict: Dict[str, List[ParsedSlot]],
                     faculty_map: Dict[str, Dict[str, List[str]]]):
        merged_value_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for rng in sheet.merged_cells.ranges:
            top_left = (rng.min_row, rng.min_col)
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_value_map[(r, c)] = top_left

        max_r = sheet.max_row
        r = 1
        current_section: Optional[str] = None

        while r <= max_r:
            c1_val = str(sheet.cell(r, 1).value or "").strip()
            c2_val = str(sheet.cell(r, 2).value or "").strip()

            sec_header = self._match_section_header(c1_val) or self._match_section_header(c2_val)
            if sec_header:
                current_section = sec_header
                if current_section not in sections_dict:
                    sections_dict[current_section] = []
                if current_section not in faculty_map:
                    faculty_map[current_section] = {}

            c1_upper = c1_val.upper()
            if c1_upper in self.DAY_NORM_MAP and current_section:
                day = self.DAY_NORM_MAP[c1_upper]
                self._parse_day_row(sheet, r, current_section, day, sheet_name,
                                    merged_value_map, all_slots, sections_dict[current_section])

            if current_section:
                for col_idx in range(1, 12):
                    c_val = str(sheet.cell(r, col_idx).value or "").strip()
                    if ":" in c_val and not any(k in c_val.upper() for k in ["DEPARTMENT", "PERIOD", "DAY", "TIME"]):
                        self._parse_faculty_legend(c_val, current_section, faculty_map)

            r += 1

    def _match_section_header(self, text: str) -> Optional[str]:
        if not text:
            return None
        upper = text.replace("\n", " ").strip().upper()
        if "DEPARTMENT" in upper or "ACADEMIC" in upper or "PERIOD" in upper or "DAY" in upper:
            return None
        if "HONORS--" in upper or "PROBABILITY" in upper or "ELEMENTARY" in upper or "INTERNSHIP" in upper:
            return None
        for prefix in self.KNOWN_SECTION_PREFIXES:
            if upper.startswith(prefix) or upper == prefix:
                clean_name = text.replace("\n", " ").strip()
                clean_name = re.sub(r"\s+", " ", clean_name)
                return clean_name
        return None

    def _get_cell_value(self, sheet: Any, r: int, c: int, merged_map: Dict[Tuple[int, int], Tuple[int, int]]) -> Any:
        if (r, c) in merged_map:
            top_r, top_c = merged_map[(r, c)]
            return sheet.cell(top_r, top_c).value
        return sheet.cell(r, c).value

    def _parse_day_row(self, sheet: Any, r: int, section: str, day: str, sheet_name: str,
                       merged_map: Dict[Tuple[int, int], Tuple[int, int]],
                       all_slots: List[ParsedSlot], section_slots: List[ParsedSlot]):
        for col, period in self.COL_PERIOD_MAP.items():
            cell_val = self._get_cell_value(sheet, r, col, merged_map)
            cell_str = str(cell_val or "").strip()

            if not cell_str or cell_str.upper() in ["BREAK", "LUNCH"]:
                continue

            subj, room, stype = self._extract_subject_room(cell_str)

            slot = ParsedSlot(
                section=section,
                day=day,
                period=period,
                subject_code=subj,
                room=room,
                subject_type=stype,
                raw_cell=cell_str,
                sheet_name=sheet_name
            )

            all_slots.append(slot)
            section_slots.append(slot)

    def _extract_subject_room(self, cell_str: str) -> Tuple[str, str, str]:
        lines = [line.strip() for line in cell_str.split("\n") if line.strip()]
        if not lines:
            return "", "", "L"

        subj = lines[0]
        room = lines[1] if len(lines) > 1 else ""

        if not room:
            parts = subj.rsplit(" ", 1)
            if len(parts) == 2 and re.match(r"^(?:[A-Z]{1,4}-?\d+|\d{3,4}[A-B]?)$", parts[1], re.IGNORECASE):
                subj = parts[0].strip()
                room = parts[1].strip()

        # Appendix B Normalizations
        # 1. Typo UFTF-13 -> AFTF-13
        if "UFTF-13" in room.upper():
            room = room.upper().replace("UFTF-13", "AFTF-13")

        # 2. AFTF-12(U-BLOCK LOCK) -> AFTF-12
        if "LOCK" in room.upper():
            room = re.sub(r"\(.*LOCK.*\)", "", room, flags=re.IGNORECASE).strip()

        # 3. External venue strings
        if ":" in room or "SEMINAR" in room.upper() or "P-BLOCK" in room.upper():
            room = "EXTERNAL"

        stype = "L"
        subj_upper = subj.upper()
        if "(P)" in subj or "LAB" in subj_upper:
            stype = "P"
        elif "(T)" in subj:
            stype = "T"
        elif "LIBRARY" in subj_upper or "LIB" == subj_upper:
            stype = "LIBRARY"
        elif "MINOR" in subj_upper or "HONOR" in subj_upper:
            stype = "MINORHONOR"
        elif any(k in subj_upper for k in ["SL/EL", "AL/IL", "SL/EL/IL", "SL_EL"]):
            stype = "SL_EL"
        elif "IDP" in subj_upper:
            stype = "IDP"
        elif "QALR" in subj_upper:
            stype = "QALR"
        elif "CRT" in subj_upper:
            stype = "CRT"

        return subj, room, stype

    def _parse_faculty_legend(self, line: str, section: str, faculty_map: Dict[str, Dict[str, List[str]]]):
        """Parse a legend line like 'Data Structures(L): Dr. Name, Dr. Name2'
        and populate faculty_map with ALL relevant normalised keys."""
        if ":" not in line:
            return
        parts = line.split(":", 1)
        subj_part = parts[0].strip()
        fac_part  = parts[1].strip()

        # CRITICAL-03 / HIGH-04: Normalise every faculty name — strip type prefixes & leading dots
        raw_fac_names = re.split(r"[,;/&]", fac_part)
        fac_names = [
            normalize_faculty_name(f)
            for f in raw_fac_names
            if f.strip() and f.strip().lower() not in ("nil", "") and not f.strip()[0].isdigit()
        ]
        fac_names = [n for n in fac_names if n]  # remove empty strings after normalisation
        if not fac_names:
            return

        if section not in faculty_map:
            faculty_map[section] = {}

        # Write raw key for backward compat
        faculty_map[section][subj_part] = fac_names

        # Derive normalised short-code and type suffix
        clean_code, suffix = normalize_subject_code(subj_part)
        if not clean_code:
            return

        # Write all variant keys so lookup by code, code+(T), code+(P) etc all hit
        faculty_map[section][clean_code] = fac_names
        faculty_map[section][f"{clean_code}(L)"] = fac_names
        faculty_map[section][f"{clean_code}(P)"] = fac_names
        faculty_map[section][f"{clean_code}(T)"] = fac_names
        faculty_map[section][f"{clean_code}(T&P)"] = fac_names
        if suffix:
            faculty_map[section][f"{clean_code}{suffix}"] = fac_names
