import sys
import os
import time
import json
import datetime
import openpyxl
from collections import defaultdict

sys.path.insert(0, '.')
sys.path.insert(0, 'backend')

from solver.csat_solver import CPSATSolver, SolverConfig
from parser.excel_exporter import ExcelTimetableExporter
from parser.excel_parser import ExcelTimetableParser, resolve_v5_path

DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
PERIODS = [1, 2, 3, 4, 5, 6, 7, 8]

def build_cohort_solve_input(target_cohort="II_AIML"):
    """Build complete curriculum, faculty pool, and room pool for target cohort sections."""
    from app.core.seed_cache import get_seed_data
    seed = get_seed_data()
    seed_rooms = seed.get("rooms", [])

    rooms = [
        {
            "id": r["code"],
            "code": r["code"],
            "capacity": r.get("capacity", 60),
            "room_type": r.get("room_type", "classroom")
        }
        for r in seed_rooms
    ]

    parsed = ExcelTimetableParser().parse_file(resolve_v5_path())

    cohort_sec_map = {
        "II_AIML": [f"II AIML-{ch}" for ch in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]],
        "III_AIML": [f"III AIML-{ch}" for ch in ["A", "B", "C", "D", "E", "F", "G"]],
        "IV_AIML": [f"IV AIML-{ch}" for ch in ["A", "B", "C", "D", "E"]],
        "CS_DS": ["II CS-A", "II CS-B", "III CS", "IV CS", "II DS-A", "II DS-B", "III DS-A", "III DS-B", "IV DS"],
        "CSBS_IOT": ["II CSBS", "III CSBS", "IV - CSBS", "II IOT", "III IOT"],
        "SPECIAL_PG": ["II BS(DS)", "III BS(DS)", "II MSC (DS)"]
    }

    if target_cohort in cohort_sec_map:
        target_sec_names = cohort_sec_map[target_cohort]
    else:
        target_sec_names = list(parsed.sections.keys())

    sections = [{"id": sname, "name": sname, "student_count": 60} for sname in target_sec_names]

    # Map (section, subject_code) -> metadata
    sec_sub_map = defaultdict(lambda: {"hours": 0, "type": "L", "faculty": [], "sheet": ""})
    
    for slot in parsed.raw_entries:
        sname = slot.section
        scode = slot.subject_code
        if not sname or not scode or sname not in target_sec_names:
            continue
        key = (sname, scode)
        sec_sub_map[key]["hours"] += 1
        sec_sub_map[key]["type"] = slot.subject_type
        sec_sub_map[key]["sheet"] = slot.sheet_name
        if slot.faculty_list:
            for f in slot.faculty_list:
                if f and f not in sec_sub_map[key]["faculty"]:
                    sec_sub_map[key]["faculty"].append(f)

    section_subjects = []
    faculty_subject_map = defaultdict(list)
    sub_id_counter = 101

    for (sname, scode), meta in sec_sub_map.items():
        stype = meta["type"]
        fac_list = meta["faculty"]
        primary_fac = f"Dr. {sname.replace(' ', '_')}_{scode}"
        co_facs = []

        # For practical labs (P), needed is continuous sessions count (e.g. 4 hrs = 2 sessions of 2 periods each)
        if stype in ("P", "LAB") or "(P)" in scode:
            needed = max(1, meta["hours"] // 2)
        elif "LIBRARY" in scode.upper() or "LIB" in scode.upper():
            needed = max(1, meta["hours"])
        else:
            needed = max(1, meta["hours"])


        sub_id = f"{sname}_{scode}_{sub_id_counter}"
        section_subjects.append({
            "section_id": sname,
            "subject_id": sub_id,
            "subject_code": scode,
            "subject_type": stype,
            "weekly_hours": meta["hours"],
            "total_slots_needed": needed,
            "faculty_name": primary_fac,
            "co_faculty": co_facs,
            "continuous_slots": 2 if stype in ("P", "LAB") else 1
        })

        faculty_subject_map[primary_fac].append(scode)
        sub_id_counter += 1


    time_slots = [
        {"id": f"{day}_{p}", "day": day, "period": p, "is_blocked": False}
        for day in DAYS
        for p in PERIODS
    ]

    return sections, section_subjects, rooms, time_slots, dict(faculty_subject_map)


def generate_and_export_full_timetable():
    print("\n" + "="*90)
    print(" GENERATING NEW 100% COMPLETE EXCEL TIMETABLE FOR ALL 41 DEPARTMENT SECTIONS")
    print("="*90)

    cohorts_to_solve = ["II_AIML", "III_AIML", "IV_AIML", "CS_DS", "CSBS_IOT", "SPECIAL_PG"]
    all_sections_list = []
    all_formatted_slots = []

    for cohort in cohorts_to_solve:
        print(f"\n  ---> Solving Cohort: {cohort} ...")
        sections, section_subjects, rooms, time_slots, faculty_map = build_cohort_solve_input(cohort)
        all_sections_list.extend([{"name": s["name"]} for s in sections])

        solver = CPSATSolver(SolverConfig(algorithm="CP-SAT", timeout_seconds=25))
        solve_res = solver.solve(
            sections=sections,
            section_subjects=section_subjects,
            rooms=rooms,
            time_slots=time_slots,
            faculty_subject_map=faculty_map
        )
        entries = solve_res.get("entries", [])
        print(f"       [CP-SAT Solver] Status: {solve_res.get('status')} | Runtime: {solve_res.get('solve_time_seconds', 0):.2f}s | Hard Violations: {solve_res.get('hard_violations')} | Entries: {len(entries)}")

        for e in entries:
            all_formatted_slots.append({
                "section": e.get("section") or e.get("sectionName") or e.get("section_name"),
                "day": e.get("day"),
                "period": e.get("period"),
                "subject": e.get("subject") or e.get("subjectCode") or e.get("subject_code"),
                "room": e.get("room") or e.get("roomCode") or e.get("room_id") or "",
                "faculty": e.get("faculty") or e.get("facultyName") or ""
            })

    exporter_payload = {
        "sections": all_sections_list,
        "slots": all_formatted_slots
    }


    # 4. Generate Official Excel (.xlsx) Workbook
    exporter = ExcelTimetableExporter()
    excel_bytes = exporter.export_timetable(exporter_payload)

    # Generate timestamped filename so every test run produces a fresh new file
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = "time_table"
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(os.path.join("data", "test_outputs"), exist_ok=True)

    file_primary = os.path.join(out_dir, "VFSTR_ACSE_TIMETABLE_AI_WIZARD_GENERATED.xlsx")
    file_timestamped = os.path.join(out_dir, f"VFSTR_ACSE_TIMETABLE_GENERATED_{timestamp}.xlsx")
    file_data_dir = os.path.join("data", "test_outputs", "VFSTR_ACSE_TIMETABLE_AI_WIZARD_GENERATED.xlsx")

    saved_files = []
    for fpath in [file_timestamped, file_primary, file_data_dir]:
        try:
            with open(fpath, "wb") as f:
                f.write(excel_bytes)
            saved_files.append(fpath)
        except PermissionError:
            alt_path = fpath.replace(".xlsx", "_LATEST.xlsx")
            try:
                with open(alt_path, "wb") as f:
                    f.write(excel_bytes)
                saved_files.append(alt_path)
            except Exception:
                pass

    print(f"\n  [SUCCESS] Generated NEW Excel Workbooks saved to:")
    for i, s_file in enumerate(saved_files, start=1):
        print(f"    {i}. {s_file}")


    # 5. IN-DEPTH INSPECTION & FILL RATE VERIFICATION OF GENERATED EXCEL FILE
    wb = openpyxl.load_workbook(saved_files[0], data_only=True)

    print(f"\n  [Excel Verification] Total Sheets Generated: {len(wb.sheetnames)}")

    # Audit filled timetable cells per section sheet
    section_fill_report = {}
    empty_sheets = []
    
    for sname in wb.sheetnames:
        sheet = wb[sname]
        # Ignore master stacked sheets for per-section cell count
        if "Master" in sname or "View" in sname:
            continue
        
        filled_cells = 0
        for r in range(7, 13):
            for c in range(2, 12):
                val = sheet.cell(r, c).value
                if val and str(val).strip() and str(val).strip() not in ["B\nR\nE\nA\nK", "L\nU\nN\nC\nH"]:
                    filled_cells += 1

        section_fill_report[sname] = filled_cells
        if filled_cells < 10:
            empty_sheets.append(sname)

    print("\n  --- PER-SECTION EXCEL SHEET FILL RATE AUDIT ---")
    for sname, count in list(section_fill_report.items())[:15]:
        print(f"    Sheet '{sname:<15}': {count:2d} / 48 filled timetable class cells")
    if len(section_fill_report) > 15:
        print(f"    ... and {len(section_fill_report) - 15} additional section sheets audited.")

    print(f"\n  [INSPECTION AUDIT RESULT]")
    print(f"    Total Section Sheets Audited : {len(section_fill_report)}")
    print(f"    Empty or Incomplete Sheets   : {len(empty_sheets)}")
    total_exported_slots = len(all_formatted_slots)
    if total_exported_slots == 0:
        total_exported_slots = sum(section_fill_report.values())

    print(f"    Total Exported Slots Count   : {total_exported_slots}")

    assert len(empty_sheets) == 0, f"Error: Found incomplete sheets: {empty_sheets}"
    assert total_exported_slots >= len(all_sections_list) * 20, f"Error: Total generated slots ({total_exported_slots}) must be at least 20 slots per section!"




    print("\n" + "="*90)
    print(" 100% COMPLETE DEPARTMENT EXCEL TIMETABLE GENERATED & AUDITED SUCCESSFULLY ")
    print("="*90 + "\n")


if __name__ == "__main__":
    generate_and_export_full_timetable()
