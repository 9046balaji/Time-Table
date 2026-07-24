import openpyxl

def inspect_minor_honors():
    wb = openpyxl.load_workbook("time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx", data_only=True)
    if "MINORHONORS" in wb.sheetnames:
        sheet = wb["MINORHONORS"]
        print("\n==================== SHEET: MINORHONORS (V5) ====================")
        for r in range(1, min(50, sheet.max_row)):
            row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, 12)]
            if any(row_vals):
                print(f"Row {r:2d}: {row_vals}")

inspect_minor_honors()
