import json
import os
import glob
import pandas as pd
import openpyxl

def audit_all():
    report = {}

    # 1. Analyze V5 Excel
    v5_path = 'time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx'
    wb_v5 = openpyxl.load_workbook(v5_path, data_only=True)
    report['v5_sheets'] = wb_v5.sheetnames
    
    # 2. Analyze 4th yr TT 17TH JULY.xlsx
    yr4_path = 'time_table/4th yr TT 17TH JULY.xlsx'
    wb_4th = openpyxl.load_workbook(yr4_path, data_only=True)
    report['4th_year_sheets'] = wb_4th.sheetnames

    # 3. Analyze JSON seed files in data/seed/
    seed_files = glob.glob('data/seed/*.json')
    report['seed_files'] = {}
    for sf in seed_files:
        filename = os.path.basename(sf)
        with open(sf, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                report['seed_files'][filename] = f"List of {len(data)} items"
            elif isinstance(data, dict):
                keys_summary = {k: (len(v) if isinstance(v, (list, dict)) else type(v).__name__) for k, v in data.items()}
                report['seed_files'][filename] = keys_summary

    # 4. Analyze generated test outputs in data/test_outputs/
    out_files = glob.glob('data/test_outputs/*.json')
    report['test_outputs'] = {}
    for of in out_files:
        filename = os.path.basename(of)
        with open(of, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                report['test_outputs'][filename] = f"List of {len(data)} items"
            elif isinstance(data, dict):
                keys_summary = {k: (len(v) if isinstance(v, (list, dict)) else type(v).__name__) for k, v in data.items()}
                report['test_outputs'][filename] = keys_summary

    print(json.dumps(report, indent=2))

if __name__ == '__main__':
    audit_all()
