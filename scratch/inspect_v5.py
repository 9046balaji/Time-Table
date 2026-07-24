import openpyxl

def inspect_workbook(filename):
    fpath = f"time_table/{filename}"
    wb = openpyxl.load_workbook(fpath, data_only=True)
    print(f"\n==================== INSPECTING: {filename} ====================")
    print("Sheets in workbook:", wb.sheetnames)

    for sname in wb.sheetnames:
        sheet = wb[sname]
        sections = []
        for r in range(1, min(300, sheet.max_row)):
            val1 = str(sheet.cell(r, 1).value or "").strip()
            val2 = str(sheet.cell(r, 2).value or "").strip()
            text = val1 if val1 else val2
            if any(k in text.upper() for k in ["AIML", "CS", "DS", "CSBS", "IOT", "BS(DS)", "MSC", "MTECH", "MINOR", "HONOR"]):
                if not any(k in text.upper() for k in ["DEPARTMENT", "PERIOD", "DAY", "ACADEMIC"]):
                    sections.append((r, text))
        print(f"Sheet [{sname}]: Found {len(sections)} section blocks -> {[s[1] for s in sections[:5]]}")

inspect_workbook("ACSE TIMETABLE (V4)  - W.e.f 14-7-2026.xlsx")
inspect_workbook("ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx")
