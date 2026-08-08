import json
import os
import openpyxl
import glob

def run_deep_analysis():
    v5_path = 'time_table/ACSE TIMETABLE (V5)  - W.e.f 15-7-2026.xlsx'
    v4_path = 'time_table/ACSE TIMETABLE (V4)  - W.e.f 14-7-2026.xlsx'
    yr4_path = 'time_table/4th yr TT 17TH JULY.xlsx'

    wb_v5 = openpyxl.load_workbook(v5_path, data_only=True)
    wb_v4 = openpyxl.load_workbook(v4_path, data_only=True)
    wb_4th = openpyxl.load_workbook(yr4_path, data_only=True)

    print("=== EXCEL SHEETS AUDIT ===")
    print(f"V5 Sheets ({len(wb_v5.sheetnames)}):", wb_v5.sheetnames)
    print(f"V4 Sheets ({len(wb_v4.sheetnames)}):", wb_v4.sheetnames)
    print(f"4th Yr Sheets ({len(wb_4th.sheetnames)}):", wb_4th.sheetnames)

    # Load seed JSONs
    with open('data/seed/original_v5_all_entries.json', 'r', encoding='utf-8') as f:
        v5_entries = json.load(f)
    with open('data/seed/original_v5_sections.json', 'r', encoding='utf-8') as f:
        v5_sections = json.load(f)
    with open('data/seed/original_v5_faculty.json', 'r', encoding='utf-8') as f:
        v5_faculty = json.load(f)
    with open('data/seed/original_v5_rooms.json', 'r', encoding='utf-8') as f:
        v5_rooms = json.load(f)
    with open('data/seed/original_v5_subjects.json', 'r', encoding='utf-8') as f:
        v5_subjects = json.load(f)

    print("\n=== SEED JSON AUDIT ===")
    print("V5 Total Slots:", len(v5_entries))
    print("V5 Total Sections:", len(v5_sections))
    print("V5 Total Faculty:", len(v5_faculty))
    print("V5 Total Rooms:", len(v5_rooms))
    print("V5 Total Subjects:", len(v5_subjects))

    # Room codes in seed JSON
    print("\nAll V5 Room Codes in JSON:", list(v5_rooms.keys()))

    # Section names in seed JSON
    print("\nAll V5 Section Names in JSON:", list(v5_sections.keys()))

    # Check 4th yr entries count
    sec_count_4th = len(wb_4th.sheetnames)
    print(f"\n4th Year Total Section Sheets: {sec_count_4th}")

if __name__ == '__main__':
    run_deep_analysis()
